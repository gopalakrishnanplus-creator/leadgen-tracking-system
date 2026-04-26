import logging
import base64
import socket
from calendar import monthrange
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo

from django.conf import settings
from django.db import connection
from django.core.mail import EmailMultiAlternatives
from django.db import transaction
from django.db.models import Avg, Count, F, Max, Q, Sum
from django.template.loader import render_to_string
from django.utils import timezone
from icalendar import Calendar, Event, vCalAddress, vText
from openpyxl import load_workbook
from sendgrid import SendGridAPIClient

from .models import (
    CallImportBatch,
    CallLog,
    CashflowImportedItem,
    CashflowPaymentPlanEntry,
    CashflowProjectedCollection,
    CashflowSnapshot,
    CashflowWorkOrderInstallment,
    CashflowWorkOrderRequest,
    ContractCollection,
    ContractCollectionContact,
    ContractCollectionFile,
    ContractCollectionInstallment,
    Meeting,
    MeetingReminder,
    Prospect,
    ProspectStatusUpdate,
    SalesConversation,
    SalesConversationBrand,
    SalesConversationContact,
    SalesConversationFile,
    SupervisorAccessEmail,
    SystemSetting,
    User,
)


logger = logging.getLogger(__name__)

MEETING_RESCHEDULE_REASON = "Meeting rescheduled"


CRM_STATUS_MAP = {
    "completed": Prospect.CRM_COMPLETED,
    "no-answer": Prospect.CRM_NO_ANSWER,
    "busy": Prospect.CRM_BUSY,
    "failed": Prospect.CRM_FAILED,
}

ACTIVE_CALLING_WORKFLOWS = {
    Prospect.WORKFLOW_READY_TO_CALL,
    Prospect.WORKFLOW_FOLLOW_UP,
}

REMINDER_GRACE_WINDOWS = {
    MeetingReminder.TYPE_WHATSAPP_INITIAL: timedelta(minutes=30),
    MeetingReminder.TYPE_WHATSAPP_PRE_MEETING: timedelta(),
    MeetingReminder.TYPE_EMAIL_DAY_BEFORE: timedelta(hours=1),
    MeetingReminder.TYPE_EMAIL_SAME_DAY: timedelta(hours=1),
    MeetingReminder.TYPE_WHATSAPP_FINAL: timedelta(minutes=15),
    MeetingReminder.TYPE_WHATSAPP_NO_SHOW: timedelta(),
    MeetingReminder.TYPE_WHATSAPP_28_DAY: timedelta(hours=24),
}

MEETING_NON_HAPPENED_STATUSES = tuple(Meeting.DID_NOT_HAPPEN_STATUSES)

REMINDER_LABELS = {
    MeetingReminder.TYPE_WHATSAPP_INITIAL: "First WhatsApp",
    MeetingReminder.TYPE_WHATSAPP_PRE_MEETING: "Pre-meeting WhatsApp",
    MeetingReminder.TYPE_EMAIL_DAY_BEFORE: "24-hour email",
    MeetingReminder.TYPE_EMAIL_SAME_DAY: "9 a.m. email",
    MeetingReminder.TYPE_WHATSAPP_FINAL: "Second WhatsApp",
    MeetingReminder.TYPE_WHATSAPP_NO_SHOW: "48-hour no-show WhatsApp",
    MeetingReminder.TYPE_WHATSAPP_28_DAY: "28-day follow-up WhatsApp",
}

CASHFLOW_CATEGORY_FILE_FIELD_MAP = {
    CashflowImportedItem.CATEGORY_PAYABLE: "payables_file",
    CashflowImportedItem.CATEGORY_PROVISION: "provisions_file",
    CashflowImportedItem.CATEGORY_RECEIVABLE: "receivables_file",
}

CASHFLOW_HEADER_ALIASES = {
    "party_name": {
        "party name",
        "ledger name",
        "account name",
        "name",
        "particulars",
        "party",
        "customer name",
        "supplier name",
    },
    "amount": {
        "amount",
        "pending amount",
        "outstanding amount",
        "closing balance",
        "balance",
        "dr amount",
        "cr amount",
    },
    "reference_number": {
        "reference",
        "ref no",
        "reference no",
        "bill ref",
        "invoice no",
        "voucher no",
        "document no",
        "bill no",
    },
    "description": {
        "description",
        "narration",
        "details",
        "particulars 2",
    },
    "due_date": {
        "due date",
        "bill due date",
        "expected date",
        "date",
    },
}


def normalize_phone(value):
    digits = phone_digits(value)
    return f"+{digits}" if digits else ""


def phone_digits(value):
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return "".join(ch for ch in format(value, "f") if ch.isdigit())
    text = str(value).strip()
    if not text:
        return ""
    try:
        decimal_value = Decimal(text.replace(",", ""))
    except (InvalidOperation, ValueError):
        decimal_value = None
    if decimal_value is not None and decimal_value == decimal_value.to_integral_value():
        return str(int(decimal_value))
    if text.startswith("+"):
        return "".join(ch for ch in text[1:] if ch.isdigit())
    return "".join(ch for ch in text if ch.isdigit())


def phone_lookup_variants(value):
    digits = phone_digits(value)
    if not digits:
        return set()
    variants = {digits, f"+{digits}"}
    if len(digits) == 10:
        variants.add(f"91{digits}")
        variants.add(f"+91{digits}")
    if len(digits) == 12 and digits.startswith("91"):
        local_digits = digits[2:]
        variants.add(local_digits)
        variants.add(f"+{local_digits}")
    return variants


def parse_report_datetime(value, tz_name):
    if not value:
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        dt = datetime.fromisoformat(str(value))
    if timezone.is_naive(dt):
        return timezone.make_aware(dt, ZoneInfo(tz_name))
    return dt.astimezone(ZoneInfo(tz_name))


def _prospect_latest_log(prospect):
    return prospect.call_logs.order_by("-started_at", "-created_at").first()


def refresh_prospect_call_metrics(prospect):
    aggregates = prospect.call_logs.aggregate(
        total_attempts=Count("id"),
        total_connected=Count("id", filter=Q(was_connected=True)),
        last_started_at=Max("started_at"),
        total_duration=Sum("duration_seconds"),
    )
    last_log = _prospect_latest_log(prospect)
    prospect.total_call_attempts = aggregates["total_attempts"] or 0
    prospect.total_connected_calls = aggregates["total_connected"] or 0
    prospect.latest_call_attempt_at = aggregates["last_started_at"]
    prospect.latest_call_duration_seconds = last_log.duration_seconds if last_log else 0
    if last_log:
        prospect.latest_crm_status = CRM_STATUS_MAP.get(last_log.crm_status, Prospect.CRM_NOT_CALLED)
    prospect.save(
        update_fields=[
            "total_call_attempts",
            "total_connected_calls",
            "latest_call_attempt_at",
            "latest_call_duration_seconds",
            "latest_crm_status",
            "updated_at",
        ]
    )


def unanswered_attempt_count(prospect):
    return prospect.call_logs.filter(crm_status=CallLog.STATUS_NO_ANSWER).count()


def refresh_prospect_import_state(prospect):
    if prospect.approval_status != Prospect.APPROVAL_ACCEPTED:
        return
    if prospect.workflow_status == Prospect.WORKFLOW_INVALID_NUMBER:
        return
    if prospect.latest_crm_status == Prospect.CRM_FAILED:
        prospect.workflow_status = Prospect.WORKFLOW_INVALID_NUMBER
        prospect.system_action_note = "Reported as an invalid number by the Exotel import."
        prospect.follow_up_date = None
        prospect.follow_up_reason = ""
        prospect.save(
            update_fields=[
                "workflow_status",
                "system_action_note",
                "follow_up_date",
                "follow_up_reason",
                "updated_at",
            ]
        )
        return
    if prospect.workflow_status not in ACTIVE_CALLING_WORKFLOWS:
        return
    unanswered_count = unanswered_attempt_count(prospect)
    if (
        prospect.latest_crm_status == Prospect.CRM_NO_ANSWER
        and unanswered_count - prospect.no_answer_reset_count >= 5
    ):
        prospect.workflow_status = Prospect.WORKFLOW_SUPERVISOR_ACTION
        prospect.system_action_note = "Attempted five times with no answer."
        prospect.save(update_fields=["workflow_status", "system_action_note", "updated_at"])


def database_healthcheck():
    with connection.cursor() as cursor:
        cursor.execute("SELECT 1")
        row = cursor.fetchone()
    return row == (1,)


def build_supervisor_report(start_date, end_date, tz_name):
    tz = ZoneInfo(tz_name)
    start_dt = timezone.make_aware(datetime.combine(start_date, datetime.min.time()), tz)
    end_dt = timezone.make_aware(datetime.combine(end_date, datetime.max.time()), tz)

    calls = CallLog.objects.filter(started_at__range=(start_dt, end_dt))
    status_updates = ProspectStatusUpdate.objects.filter(created_at__range=(start_dt, end_dt))
    scheduled_updates = status_updates.filter(outcome=ProspectStatusUpdate.OUTCOME_SCHEDULED)
    new_scheduled_updates = scheduled_updates.exclude(reason=MEETING_RESCHEDULE_REASON)
    rescheduled_updates = scheduled_updates.filter(reason=MEETING_RESCHEDULE_REASON)
    meetings_outcome = Meeting.objects.filter(outcome_updated_at__range=(start_dt, end_dt))
    imports = CallImportBatch.objects.filter(import_date__range=(start_date, end_date))

    summary = {
        "attempts": calls.count(),
        "connects": calls.filter(was_connected=True).count(),
        "connect_rate": 0,
        "avg_connected_duration": calls.filter(was_connected=True).aggregate(avg=Avg("duration_seconds"))["avg"] or 0,
        "no_answer_count": calls.filter(crm_status=CallLog.STATUS_NO_ANSWER).count(),
        "busy_count": calls.filter(crm_status=CallLog.STATUS_BUSY).count(),
        "failed_count": calls.filter(crm_status=CallLog.STATUS_FAILED).count(),
        "follow_ups": status_updates.filter(outcome=ProspectStatusUpdate.OUTCOME_FOLLOW_UP).count(),
        "declines": status_updates.filter(outcome=ProspectStatusUpdate.OUTCOME_DECLINED).count(),
        "meetings_scheduled": new_scheduled_updates.count(),
        "meetings_rescheduled": rescheduled_updates.count(),
        "meetings_happened": meetings_outcome.filter(status=Meeting.STATUS_HAPPENED).count(),
        "meetings_not_happened": meetings_outcome.filter(status__in=MEETING_NON_HAPPENED_STATUSES).count(),
        "imports_count": imports.count(),
        "imported_rows": imports.aggregate(total=Sum("total_rows"))["total"] or 0,
        "unmatched_import_rows": imports.aggregate(total=Sum("unmatched_rows"))["total"] or 0,
        "duplicate_import_rows": imports.aggregate(total=Sum("duplicate_rows"))["total"] or 0,
    }
    if summary["attempts"]:
        summary["connect_rate"] = round((summary["connects"] / summary["attempts"]) * 100, 1)

    staff_metrics = []
    for staff in User.objects.filter(role=User.ROLE_STAFF).order_by("name"):
        staff_calls = calls.filter(staff=staff)
        attempts = staff_calls.count()
        connects = staff_calls.filter(was_connected=True).count()
        scheduled = new_scheduled_updates.filter(staff=staff).count()
        rescheduled = rescheduled_updates.filter(staff=staff).count()
        happened = meetings_outcome.filter(scheduled_by=staff, status=Meeting.STATUS_HAPPENED).count()
        not_happened = meetings_outcome.filter(scheduled_by=staff, status__in=MEETING_NON_HAPPENED_STATUSES).count()
        follow_ups = status_updates.filter(staff=staff, outcome=ProspectStatusUpdate.OUTCOME_FOLLOW_UP).count()
        declines = status_updates.filter(staff=staff, outcome=ProspectStatusUpdate.OUTCOME_DECLINED).count()
        staff_metrics.append(
            {
                "staff": staff,
                "attempts": attempts,
                "connects": connects,
                "connect_rate": round((connects / attempts) * 100, 1) if attempts else 0,
                "follow_ups": follow_ups,
                "declines": declines,
                "meetings_scheduled": scheduled,
                "meetings_rescheduled": rescheduled,
                "meetings_happened": happened,
                "meetings_not_happened": not_happened,
            }
        )

    import_batches = list(imports.order_by("-import_date", "-created_at")[:10])
    return {
        "summary": summary,
        "staff_metrics": staff_metrics,
        "import_batches": import_batches,
        "start_dt": start_dt,
        "end_dt": end_dt,
    }


def build_daily_target_report(target_date, tz_name):
    tz = ZoneInfo(tz_name)
    start_dt = timezone.make_aware(datetime.combine(target_date, datetime.min.time()), tz)
    end_dt = timezone.make_aware(datetime.combine(target_date, datetime.max.time()), tz)
    staff_rows = []

    for staff in User.objects.filter(role=User.ROLE_STAFF, is_active=True).order_by("name", "email"):
        assigned_prospects = Prospect.objects.filter(
            assigned_to=staff,
            approval_status=Prospect.APPROVAL_ACCEPTED,
            created_at__lte=end_dt,
        ).filter(Q(accepted_at__isnull=True) | Q(accepted_at__lte=end_dt))
        target_queryset = assigned_prospects.filter(
            Q(workflow_status__in=ACTIVE_CALLING_WORKFLOWS)
            | Q(call_logs__started_at__range=(start_dt, end_dt))
        ).distinct()
        status_updates = ProspectStatusUpdate.objects.filter(
            staff=staff,
            created_at__range=(start_dt, end_dt),
        )
        actual_attempts = CallLog.objects.filter(staff=staff, started_at__range=(start_dt, end_dt)).count()
        follow_up_updates = status_updates.filter(outcome=ProspectStatusUpdate.OUTCOME_FOLLOW_UP).count()
        declined_updates = status_updates.filter(outcome=ProspectStatusUpdate.OUTCOME_DECLINED).count()
        scheduled_updates = status_updates.filter(outcome=ProspectStatusUpdate.OUTCOME_SCHEDULED).count()
        staff_rows.append(
            {
                "staff": staff,
                "target_count": target_queryset.count(),
                "actual_attempts": actual_attempts,
                "delta": actual_attempts - target_queryset.count(),
                "follow_up_updates": follow_up_updates,
                "declined_updates": declined_updates,
                "scheduled_updates": scheduled_updates,
            }
        )

    return {
        "target_date": target_date,
        "staff_rows": staff_rows,
    }


def import_exotel_report(batch):
    workbook = load_workbook(batch.uploaded_file.path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [str(value).strip() for value in next(sheet.iter_rows(values_only=True))]
    required = ["Call SID", "Start Time", "End Time", "From", "To", "Direction", "Status"]
    missing = [col for col in required if col not in header]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    indexes = {name: header.index(name) for name in required}
    touched_prospects = set()
    tz_name = SystemSetting.load().default_timezone

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        batch.total_rows += 1
        row_data = {col: row[idx] for col, idx in indexes.items()}
        call_sid = str(row_data["Call SID"]).strip()
        from_number = normalize_phone(row_data["From"])
        to_number = normalize_phone(row_data["To"])
        crm_status = str(row_data["Status"]).strip().lower()
        if crm_status not in dict(CallLog.STATUS_CHOICES):
            crm_status = CallLog.STATUS_FAILED

        staff = User.objects.filter(
            role=User.ROLE_STAFF,
            calling_number__in=phone_lookup_variants(row_data["From"]),
        ).first()
        prospect = None
        if staff:
            prospect = Prospect.objects.filter(
                assigned_to=staff,
                phone_number__in=phone_lookup_variants(row_data["To"]),
            ).first()
        matched = bool(staff and prospect)
        defaults = {
            "batch": batch,
            "staff": staff,
            "prospect": prospect,
            "started_at": parse_report_datetime(row_data["Start Time"], tz_name),
            "ended_at": parse_report_datetime(row_data["End Time"], tz_name),
            "from_number": from_number,
            "to_number": to_number,
            "direction": str(row_data["Direction"]).strip().lower(),
            "crm_status": crm_status,
            "was_connected": crm_status == CallLog.STATUS_COMPLETED,
            "matched": matched,
            "raw_data": {k: str(v) if v is not None else "" for k, v in row_data.items()},
        }
        if defaults["started_at"] and defaults["ended_at"]:
            defaults["duration_seconds"] = max(
                0,
                int((defaults["ended_at"] - defaults["started_at"]).total_seconds()),
            )
        else:
            defaults["duration_seconds"] = 0
        _, created = CallLog.objects.update_or_create(call_sid=call_sid, defaults=defaults)
        if created:
            if matched:
                batch.matched_rows += 1
            else:
                batch.unmatched_rows += 1
        else:
            batch.duplicate_rows += 1
        if prospect:
            touched_prospects.add(prospect.pk)

    batch.save(
        update_fields=[
            "total_rows",
            "matched_rows",
            "unmatched_rows",
            "duplicate_rows",
        ]
    )
    for prospect in Prospect.objects.filter(pk__in=touched_prospects):
        refresh_prospect_call_metrics(prospect)
        refresh_prospect_import_state(prospect)


def unique_emails(*groups):
    seen = set()
    emails = []
    for group in groups:
        for email in group:
            normalized = (email or "").strip().lower()
            if normalized and normalized not in seen:
                seen.add(normalized)
                emails.append(normalized)
    return emails


def mask_secret(value):
    if not value:
        return ""
    if len(value) <= 8:
        return "*" * len(value)
    return f"{value[:4]}...{value[-4:]}"


def probe_sendgrid_connectivity():
    try:
        addresses = socket.getaddrinfo("api.sendgrid.com", 443, type=socket.SOCK_STREAM)
        hosts = sorted({item[4][0] for item in addresses if item[4]})
        return {"dns_ok": True, "resolved_hosts": hosts[:5]}
    except Exception as exc:
        return {
            "dns_ok": False,
            "error_type": exc.__class__.__name__,
            "error_message": str(exc),
        }


def build_calendar_invite(meeting, settings_obj):
    tz = ZoneInfo(settings_obj.default_timezone)
    start = meeting.scheduled_for.astimezone(tz)
    end = start + timedelta(minutes=30)
    organizer_email = (settings.DEFAULT_FROM_EMAIL or settings_obj.supervisor_sender_email or "").strip().lower()

    calendar = Calendar()
    calendar.add("prodid", "-//Leadgen Tracking System//inditech.co.in//")
    calendar.add("version", "2.0")
    calendar.add("method", "REQUEST")

    event = Event()
    event.add("uid", f"meeting-{meeting.pk}@leadgen.local")
    event.add("summary", f"Meeting with {meeting.prospect.contact_name} - {meeting.prospect.company_name}")
    event.add("dtstart", start)
    event.add("dtend", end)
    event.add("dtstamp", timezone.now())
    if meeting.meeting_link:
        event.add("location", meeting.meeting_link)
        event.add("url", meeting.meeting_link)
    description_lines = [
        f"Prospect: {meeting.prospect.contact_name}",
        f"Company: {meeting.prospect.company_name}",
        f"LinkedIn: {meeting.prospect.linkedin_url}",
        f"Lead Gen Staff: {meeting.scheduled_by.name}",
        f"Meeting Platform: {meeting.get_meeting_platform_display()}",
    ]
    description_lines.extend(meeting.meeting_access_lines)
    event.add(
        "description",
        "\n".join([line for line in description_lines if line]),
    )
    organizer = vCalAddress(f"MAILTO:{organizer_email}")
    organizer.params["cn"] = vText(settings_obj.supervisor_name)
    organizer.params["role"] = vText("CHAIR")
    event["organizer"] = organizer
    for email in meeting.recipient_emails:
        normalized_email = (email or "").strip().lower()
        if not normalized_email or normalized_email == organizer_email:
            continue
        attendee = vCalAddress(f"MAILTO:{email}")
        attendee.params["cn"] = vText(email)
        attendee.params["role"] = vText("REQ-PARTICIPANT")
        event.add("attendee", attendee, encode=0)
    calendar.add_component(event)
    return calendar.to_ical()


def _deliver_email(subject, html_body, text_body, to_emails, cc_emails=None, attachments=None):
    cc_emails = cc_emails or []
    attachments = attachments or []
    if settings.SENDGRID_API_KEY:
        payload = {
            "personalizations": [
                {
                    "to": [{"email": email} for email in to_emails],
                    "subject": subject,
                }
            ],
            "from": {"email": settings.DEFAULT_FROM_EMAIL},
            "reply_to": {"email": settings.REPLY_TO_EMAIL},
            "content": [
                {"type": "text/plain", "value": text_body},
                {"type": "text/html", "value": html_body},
            ],
        }
        if cc_emails:
            payload["personalizations"][0]["cc"] = [{"email": email} for email in cc_emails]
        if attachments:
            payload["attachments"] = [
                {
                    "content": base64.b64encode(attachment["content"]).decode("utf-8"),
                    "filename": attachment["filename"],
                    "type": attachment["type"],
                    "disposition": "attachment",
                }
                for attachment in attachments
            ]
        response = SendGridAPIClient(settings.SENDGRID_API_KEY).client.mail.send.post(request_body=payload)
        response_body = getattr(response, "body", b"")
        if isinstance(response_body, bytes):
            response_body = response_body.decode("utf-8", errors="ignore")
        if response.status_code < 200 or response.status_code >= 300:
            raise RuntimeError(f"SendGrid returned status {response.status_code}: {response_body}")
        return {
            "transport": "sendgrid",
            "status_code": response.status_code,
            "response_body": response_body[:1000],
            "payload_summary": {
                "from_email": settings.DEFAULT_FROM_EMAIL,
                "reply_to": settings.REPLY_TO_EMAIL,
                "to_count": len(to_emails),
                "cc_count": len(cc_emails),
                "attachment_count": len(attachments),
            },
        }

    email = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=to_emails,
        cc=cc_emails,
        reply_to=[settings.REPLY_TO_EMAIL],
    )
    email.attach_alternative(html_body, "text/html")
    for attachment in attachments:
        email.attach(attachment["filename"], attachment["content"], attachment["type"])
    sent_count = email.send()
    return {
        "transport": "django-email-backend",
        "email_backend": settings.EMAIL_BACKEND,
        "sent_count": sent_count,
        "payload_summary": {
            "from_email": settings.DEFAULT_FROM_EMAIL,
            "reply_to": settings.REPLY_TO_EMAIL,
            "to_count": len(to_emails),
            "cc_count": len(cc_emails),
            "attachment_count": len(attachments),
        },
    }


def send_email(subject, html_body, text_body, to_emails, cc_emails=None, attachments=None):
    _deliver_email(
        subject=subject,
        html_body=html_body,
        text_body=text_body,
        to_emails=to_emails,
        cc_emails=cc_emails,
        attachments=attachments,
    )


def send_test_email_diagnostic(recipient_email):
    timestamp = timezone.now()
    subject = f"Leadgen email diagnostics {timestamp:%Y-%m-%d %H:%M:%S %Z}"
    html_body = (
        "<p>This is a diagnostic test email from the Leadgen Tracking System.</p>"
        f"<p>Timestamp: {timestamp.isoformat()}</p>"
        f"<p>Recipient: {recipient_email}</p>"
        f"<p>From: {settings.DEFAULT_FROM_EMAIL}</p>"
        f"<p>Reply-To: {settings.REPLY_TO_EMAIL}</p>"
    )
    text_body = (
        "This is a diagnostic test email from the Leadgen Tracking System.\n"
        f"Timestamp: {timestamp.isoformat()}\n"
        f"Recipient: {recipient_email}\n"
        f"From: {settings.DEFAULT_FROM_EMAIL}\n"
        f"Reply-To: {settings.REPLY_TO_EMAIL}\n"
    )
    diagnostics = {
        "success": False,
        "timestamp": timestamp.isoformat(),
        "app_env": settings.APP_ENV if hasattr(settings, "APP_ENV") else "",
        "email_backend": settings.EMAIL_BACKEND,
        "sendgrid_configured": bool(settings.SENDGRID_API_KEY),
        "sendgrid_api_key_masked": mask_secret(settings.SENDGRID_API_KEY),
        "from_email": settings.DEFAULT_FROM_EMAIL,
        "reply_to_email": settings.REPLY_TO_EMAIL,
        "to_email": recipient_email,
        "sendgrid_connectivity": probe_sendgrid_connectivity(),
    }
    try:
        delivery = _deliver_email(
            subject=subject,
            html_body=html_body,
            text_body=text_body,
            to_emails=[recipient_email],
        )
        diagnostics["success"] = True
        diagnostics["delivery"] = delivery
    except Exception as exc:
        diagnostics["error_type"] = exc.__class__.__name__
        diagnostics["error_message"] = str(exc)
        logger.exception("Email diagnostics failed for recipient=%s", recipient_email)
    return diagnostics


def send_meeting_invitation(meeting):
    settings_obj = SystemSetting.load()
    html_body = render_to_string(
        "emails/meeting_invite.html",
        {"meeting": meeting, "settings_obj": settings_obj},
    )
    text_body = render_to_string(
        "emails/meeting_invite.txt",
        {"meeting": meeting, "settings_obj": settings_obj},
    )
    send_email(
        subject=f"20-min: Outcome-Linked Campaign – Discussion with {meeting.prospect.company_name}",
        html_body=html_body,
        text_body=text_body,
        to_emails=[meeting.prospect_email],
        cc_emails=[email for email in meeting.recipient_emails if email != meeting.prospect_email],
        attachments=[
            {
                "filename": "meeting-invite.ics",
                "content": build_calendar_invite(meeting, settings_obj),
                "type": "text/calendar",
            }
        ],
    )
    meeting.mark_invite_sent()


def send_did_not_happen_email(meeting):
    settings_obj = SystemSetting.load()
    html_body = render_to_string(
        "emails/meeting_did_not_happen.html",
        {"meeting": meeting, "settings_obj": settings_obj},
    )
    text_body = render_to_string(
        "emails/meeting_did_not_happen.txt",
        {"meeting": meeting, "settings_obj": settings_obj},
    )
    send_email(
        subject=f"Meeting did not happen: {meeting.prospect.company_name} / {meeting.prospect.contact_name}",
        html_body=html_body,
        text_body=text_body,
        to_emails=[meeting.scheduled_by.email],
        cc_emails=[settings_obj.supervisor_sender_email],
    )


def meeting_reminder_due_at(meeting, reminder_type, tz_name):
    tz = ZoneInfo(tz_name)
    scheduled_local = meeting.scheduled_for.astimezone(tz)
    created_local = meeting.created_at.astimezone(tz)
    if reminder_type == MeetingReminder.TYPE_WHATSAPP_INITIAL:
        return created_local
    if reminder_type == MeetingReminder.TYPE_WHATSAPP_PRE_MEETING:
        return created_local + timedelta(hours=24)
    if reminder_type == MeetingReminder.TYPE_EMAIL_DAY_BEFORE:
        if timedelta(hours=1) < (scheduled_local - created_local) <= timedelta(hours=24):
            return created_local
        return scheduled_local - timedelta(hours=24)
    if reminder_type == MeetingReminder.TYPE_EMAIL_SAME_DAY:
        return timezone.make_aware(
            datetime.combine(scheduled_local.date(), datetime.min.time()).replace(hour=9),
            tz,
        )
    if reminder_type == MeetingReminder.TYPE_WHATSAPP_FINAL:
        return scheduled_local - timedelta(hours=1)
    if reminder_type == MeetingReminder.TYPE_WHATSAPP_NO_SHOW:
        return scheduled_local + timedelta(hours=48)
    if reminder_type == MeetingReminder.TYPE_WHATSAPP_28_DAY:
        return scheduled_local + timedelta(days=28)
    raise ValueError(f"Unsupported reminder type: {reminder_type}")


def reminder_log_map(meeting):
    return {reminder.reminder_type: reminder for reminder in meeting.reminders.all()}


def meeting_has_later_follow_up(meeting):
    return Meeting.objects.filter(
        prospect=meeting.prospect,
        scheduled_for__gt=meeting.scheduled_for,
    ).exclude(pk=meeting.pk).exists()


def reminder_is_applicable(meeting, reminder_type):
    if reminder_type in {
        MeetingReminder.TYPE_WHATSAPP_INITIAL,
        MeetingReminder.TYPE_WHATSAPP_PRE_MEETING,
        MeetingReminder.TYPE_EMAIL_DAY_BEFORE,
        MeetingReminder.TYPE_EMAIL_SAME_DAY,
        MeetingReminder.TYPE_WHATSAPP_FINAL,
    }:
        return meeting.status in {
            Meeting.STATUS_SCHEDULED,
            Meeting.STATUS_HAPPENED,
            Meeting.STATUS_DID_NOT_HAPPEN,
            Meeting.STATUS_NO_SHOW,
        }
    if reminder_type == MeetingReminder.TYPE_WHATSAPP_NO_SHOW:
        return meeting.status == Meeting.STATUS_NO_SHOW
    if reminder_type == MeetingReminder.TYPE_WHATSAPP_28_DAY:
        return meeting.status in MEETING_NON_HAPPENED_STATUSES and not meeting_has_later_follow_up(meeting)
    return False


def available_whatsapp_reminder_choices(meeting):
    choice_map = {
        MeetingReminder.TYPE_WHATSAPP_INITIAL: "First WhatsApp reminder",
        MeetingReminder.TYPE_WHATSAPP_PRE_MEETING: "Pre-meeting WhatsApp reminder",
        MeetingReminder.TYPE_WHATSAPP_FINAL: "Second WhatsApp reminder",
        MeetingReminder.TYPE_WHATSAPP_NO_SHOW: "48-hour no-show WhatsApp",
        MeetingReminder.TYPE_WHATSAPP_28_DAY: "28-day follow-up WhatsApp",
    }
    ordered_types = [
        MeetingReminder.TYPE_WHATSAPP_INITIAL,
        MeetingReminder.TYPE_WHATSAPP_PRE_MEETING,
        MeetingReminder.TYPE_WHATSAPP_FINAL,
        MeetingReminder.TYPE_WHATSAPP_NO_SHOW,
        MeetingReminder.TYPE_WHATSAPP_28_DAY,
    ]
    return [
        (reminder_type, choice_map[reminder_type])
        for reminder_type in ordered_types
        if reminder_is_applicable(meeting, reminder_type)
    ]


def reminder_status_for_meeting(meeting, reminder_type, tz_name, now=None):
    now = now or timezone.now()
    if not reminder_is_applicable(meeting, reminder_type):
        return {
            "due_at": None,
            "reminder": None,
            "is_sent": False,
            "is_missed": False,
            "is_applicable": False,
        }
    due_at = meeting_reminder_due_at(meeting, reminder_type, tz_name)
    reminder = reminder_log_map(meeting).get(reminder_type)
    if reminder:
        return {
            "due_at": due_at,
            "reminder": reminder,
            "is_sent": True,
            "is_missed": False,
            "is_applicable": True,
        }
    grace = REMINDER_GRACE_WINDOWS[reminder_type]
    return {
        "due_at": due_at,
        "reminder": None,
        "is_sent": False,
        "is_missed": now >= due_at + grace,
        "is_applicable": True,
    }


def log_whatsapp_reminder(meeting, reminder_type, recipient_number, screenshot, sent_by):
    defaults = {
        "recipient_number": normalize_phone(recipient_number) or recipient_number.strip(),
        "screenshot": screenshot,
        "sent_by": sent_by,
        "sent_at": timezone.now(),
    }
    reminder, _ = MeetingReminder.objects.update_or_create(
        meeting=meeting,
        reminder_type=reminder_type,
        defaults=defaults,
    )
    return reminder


def send_meeting_reminder_email(meeting, reminder_type):
    settings_obj = SystemSetting.load()
    if reminder_type == MeetingReminder.TYPE_EMAIL_DAY_BEFORE:
        html_template = "emails/meeting_reminder_day_before.html"
        text_template = "emails/meeting_reminder_day_before.txt"
    elif reminder_type == MeetingReminder.TYPE_EMAIL_SAME_DAY:
        html_template = "emails/meeting_reminder_same_day.html"
        text_template = "emails/meeting_reminder_same_day.txt"
    else:
        raise ValueError(f"Unsupported email reminder type: {reminder_type}")

    html_body = render_to_string(
        html_template,
        {"meeting": meeting, "settings_obj": settings_obj},
    )
    text_body = render_to_string(
        text_template,
        {"meeting": meeting, "settings_obj": settings_obj},
    )
    recipients = unique_emails(
        [meeting.prospect_email],
        active_leadgen_supervisor_emails(),
    )
    send_email(
        subject=f"20-min: Outcome-Linked Campaign - Discussion with {meeting.prospect.company_name}",
        html_body=html_body,
        text_body=text_body,
        to_emails=[recipients[0]],
        cc_emails=recipients[1:],
    )
    return MeetingReminder.objects.create(
        meeting=meeting,
        reminder_type=reminder_type,
        recipient_number="",
        sent_at=timezone.now(),
    )


def send_due_meeting_reminder_emails(now=None):
    now = now or timezone.now()
    settings_obj = SystemSetting.load()
    reminders_sent = 0
    meetings = Meeting.objects.filter(status=Meeting.STATUS_SCHEDULED).select_related("prospect", "scheduled_by")
    for meeting in meetings:
        existing = reminder_log_map(meeting)
        for reminder_type in (
            MeetingReminder.TYPE_EMAIL_DAY_BEFORE,
            MeetingReminder.TYPE_EMAIL_SAME_DAY,
        ):
            if reminder_type in existing:
                continue
            if meeting_reminder_due_at(meeting, reminder_type, settings_obj.default_timezone) <= now:
                try:
                    send_meeting_reminder_email(meeting, reminder_type)
                except Exception:
                    logger.exception(
                        "Failed to send meeting reminder email for meeting_id=%s reminder_type=%s",
                        meeting.pk,
                        reminder_type,
                    )
                    continue
                reminders_sent += 1
    return reminders_sent


def _send_immediate_day_before_reminder_after_commit(meeting_id):
    try:
        meeting = Meeting.objects.select_related("prospect", "scheduled_by").prefetch_related("reminders").get(pk=meeting_id)
        if meeting.status != Meeting.STATUS_SCHEDULED:
            return
        settings_obj = SystemSetting.load()
        existing = reminder_log_map(meeting)
        if MeetingReminder.TYPE_EMAIL_DAY_BEFORE in existing:
            return
        tz = ZoneInfo(settings_obj.default_timezone)
        now_local = timezone.localtime(timezone.now(), tz)
        scheduled_local = meeting.scheduled_for.astimezone(tz)
        time_until_meeting = scheduled_local - now_local
        if not (timedelta(hours=1) < time_until_meeting <= timedelta(hours=24)):
            return
        send_meeting_reminder_email(meeting, MeetingReminder.TYPE_EMAIL_DAY_BEFORE)
    except Exception:
        logger.exception(
            "Failed to send immediate 24-hour reminder email for meeting_id=%s",
            meeting_id,
        )


def build_reminder_dashboard(tz_name, now=None):
    now = now or timezone.now()
    recent_happened_cutoff = now - timedelta(days=10)
    recent_non_happened_cutoff = now - timedelta(days=35)
    meetings = (
        Meeting.objects.select_related("prospect", "scheduled_by")
        .prefetch_related("reminders")
        .filter(
            Q(status=Meeting.STATUS_SCHEDULED)
            | Q(status=Meeting.STATUS_HAPPENED, outcome_updated_at__gte=recent_happened_cutoff)
            | Q(status__in=MEETING_NON_HAPPENED_STATUSES, outcome_updated_at__gte=recent_non_happened_cutoff)
        )
        .order_by("scheduled_for")
    )
    rows = []
    summary = {
        "first_whatsapp_sent": 0,
        "pre_meeting_whatsapp_sent": 0,
        "second_whatsapp_sent": 0,
        "no_show_whatsapp_sent": 0,
        "twenty_eight_day_whatsapp_sent": 0,
        "missed_total": 0,
    }
    for meeting in meetings:
        first_whatsapp = reminder_status_for_meeting(meeting, MeetingReminder.TYPE_WHATSAPP_INITIAL, tz_name, now=now)
        pre_meeting_whatsapp = reminder_status_for_meeting(
            meeting,
            MeetingReminder.TYPE_WHATSAPP_PRE_MEETING,
            tz_name,
            now=now,
        )
        day_before_email = reminder_status_for_meeting(meeting, MeetingReminder.TYPE_EMAIL_DAY_BEFORE, tz_name, now=now)
        same_day_email = reminder_status_for_meeting(meeting, MeetingReminder.TYPE_EMAIL_SAME_DAY, tz_name, now=now)
        final_whatsapp = reminder_status_for_meeting(meeting, MeetingReminder.TYPE_WHATSAPP_FINAL, tz_name, now=now)
        no_show_whatsapp = reminder_status_for_meeting(
            meeting,
            MeetingReminder.TYPE_WHATSAPP_NO_SHOW,
            tz_name,
            now=now,
        )
        twenty_eight_day_whatsapp = reminder_status_for_meeting(
            meeting,
            MeetingReminder.TYPE_WHATSAPP_28_DAY,
            tz_name,
            now=now,
        )
        missed = [
            REMINDER_LABELS[reminder_type]
            for reminder_type, status in (
                (MeetingReminder.TYPE_WHATSAPP_INITIAL, first_whatsapp),
                (MeetingReminder.TYPE_WHATSAPP_PRE_MEETING, pre_meeting_whatsapp),
                (MeetingReminder.TYPE_EMAIL_DAY_BEFORE, day_before_email),
                (MeetingReminder.TYPE_EMAIL_SAME_DAY, same_day_email),
                (MeetingReminder.TYPE_WHATSAPP_FINAL, final_whatsapp),
                (MeetingReminder.TYPE_WHATSAPP_NO_SHOW, no_show_whatsapp),
                (MeetingReminder.TYPE_WHATSAPP_28_DAY, twenty_eight_day_whatsapp),
            )
            if status["is_missed"]
        ]
        if first_whatsapp["is_sent"]:
            summary["first_whatsapp_sent"] += 1
        if pre_meeting_whatsapp["is_sent"]:
            summary["pre_meeting_whatsapp_sent"] += 1
        if final_whatsapp["is_sent"]:
            summary["second_whatsapp_sent"] += 1
        if no_show_whatsapp["is_sent"]:
            summary["no_show_whatsapp_sent"] += 1
        if twenty_eight_day_whatsapp["is_sent"]:
            summary["twenty_eight_day_whatsapp_sent"] += 1
        summary["missed_total"] += len(missed)
        rows.append(
            {
                "meeting": meeting,
                "first_whatsapp": first_whatsapp,
                "pre_meeting_whatsapp": pre_meeting_whatsapp,
                "day_before_email": day_before_email,
                "same_day_email": same_day_email,
                "final_whatsapp": final_whatsapp,
                "no_show_whatsapp": no_show_whatsapp,
                "twenty_eight_day_whatsapp": twenty_eight_day_whatsapp,
                "missed": missed,
            }
        )
    return {
        "summary": summary,
        "rows": rows,
        "recent_happened_cutoff": recent_happened_cutoff,
        "recent_non_happened_cutoff": recent_non_happened_cutoff,
    }


def active_sales_manager_emails():
    return list(
        User.objects.filter(role=User.ROLE_SALES_MANAGER, is_active=True)
        .exclude(email="")
        .values_list("email", flat=True)
    )


def default_sales_manager():
    managers = list(User.objects.filter(role=User.ROLE_SALES_MANAGER, is_active=True).order_by("name", "email")[:2])
    if len(managers) == 1:
        return managers[0]
    return None


def business_localdate(tz_name=None, now=None):
    tz = ZoneInfo(tz_name or SystemSetting.load().default_timezone)
    current_time = now or timezone.now()
    return timezone.localtime(current_time, tz).date()


def active_finance_manager_emails():
    return list(
        User.objects.filter(role=User.ROLE_FINANCE_MANAGER, is_active=True)
        .exclude(email="")
        .values_list("email", flat=True)
    )


def active_leadgen_supervisor_emails():
    return list(
        SupervisorAccessEmail.objects.filter(
            access_level=SupervisorAccessEmail.ACCESS_LEADGEN_SUPERVISOR,
            is_active=True,
        )
        .exclude(email="")
        .values_list("email", flat=True)
    )


def active_business_manager_emails():
    return list(
        User.objects.filter(role=User.ROLE_BUSINESS_MANAGER, is_active=True)
        .exclude(email="")
        .values_list("email", flat=True)
    )


def latest_cashflow_snapshot():
    return CashflowSnapshot.objects.order_by("-as_of_date", "-created_at").first()


def finance_upload_complete_for_date(target_date=None):
    return CashflowSnapshot.objects.filter(as_of_date=target_date or business_localdate()).exists()


def normalize_cashflow_header(value):
    return " ".join("".join(ch.lower() if ch.isalnum() else " " for ch in str(value or "")).split())


def parse_cashflow_date(value, tz_name=None):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return timezone.localtime(value, ZoneInfo(tz_name or SystemSetting.load().default_timezone)).date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def parse_cashflow_amount(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    text = text.replace("(", "-").replace(")", "")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def resolve_cashflow_columns(header_row):
    normalized = {normalize_cashflow_header(value): index for index, value in enumerate(header_row)}
    resolved = {}
    for field_name, aliases in CASHFLOW_HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                resolved[field_name] = normalized[alias]
                break
    missing = [field for field in ("party_name", "amount") if field not in resolved]
    if missing:
        raise ValueError(
            "Could not find the required Tally columns: "
            + ", ".join(missing)
            + ". Accepted header names are flexible, but the file must contain a party/entity name and amount column."
        )
    return resolved


def build_cashflow_source_key(category, party_name, reference_number, description, due_date, duplicate_index=1):
    normalized_bits = [
        category,
        normalize_cashflow_header(party_name),
        normalize_cashflow_header(reference_number),
        normalize_cashflow_header(description),
        due_date.isoformat() if due_date else "",
    ]
    key = "|".join(normalized_bits)
    if duplicate_index > 1:
        key = f"{key}|{duplicate_index}"
    return key


def import_cashflow_snapshot(snapshot):
    imported_counts = {}
    for category, file_field in CASHFLOW_CATEGORY_FILE_FIELD_MAP.items():
        imported_counts[category] = _import_cashflow_file(
            snapshot=snapshot,
            category=category,
            file_path=getattr(snapshot, file_field).path,
        )
    return imported_counts


def _import_cashflow_file(snapshot, category, file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = list(next(sheet.iter_rows(values_only=True)))
    column_map = resolve_cashflow_columns(header)

    CashflowImportedItem.objects.filter(category=category, is_current=True).update(is_current=False)
    duplicate_counts = {}
    touched_ids = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        party_name = str(row[column_map["party_name"]] or "").strip()
        amount = parse_cashflow_amount(row[column_map["amount"]])
        if not party_name or amount is None:
            continue
        reference_number = ""
        if "reference_number" in column_map:
            reference_number = str(row[column_map["reference_number"]] or "").strip()
        description = ""
        if "description" in column_map:
            description = str(row[column_map["description"]] or "").strip()
        due_date = None
        if "due_date" in column_map:
            due_date = parse_cashflow_date(row[column_map["due_date"]], tz_name=SystemSetting.load().default_timezone)
        duplicate_identity = (
            normalize_cashflow_header(party_name),
            normalize_cashflow_header(reference_number),
            normalize_cashflow_header(description),
            due_date.isoformat() if due_date else "",
        )
        duplicate_counts[duplicate_identity] = duplicate_counts.get(duplicate_identity, 0) + 1
        source_key = build_cashflow_source_key(
            category=category,
            party_name=party_name,
            reference_number=reference_number,
            description=description,
            due_date=due_date,
            duplicate_index=duplicate_counts[duplicate_identity],
        )
        item, _ = CashflowImportedItem.objects.update_or_create(
            category=category,
            source_key=source_key,
            defaults={
                "snapshot": snapshot,
                "party_name": party_name,
                "description": description,
                "reference_number": reference_number,
                "amount": amount,
                "due_date": due_date,
                "raw_row": {
                    str(header[index]): ("" if value is None else str(value))
                    for index, value in enumerate(row)
                },
                "is_current": True,
            },
        )
        touched_ids.append(item.pk)
    if touched_ids:
        CashflowImportedItem.objects.filter(pk__in=touched_ids).update(is_current=True, snapshot=snapshot)
    return len(touched_ids)


def current_cashflow_items():
    return CashflowImportedItem.objects.filter(is_current=True)


def current_cashflow_outflow_items():
    return current_cashflow_items().filter(
        category__in=[CashflowImportedItem.CATEGORY_PAYABLE, CashflowImportedItem.CATEGORY_PROVISION]
    ).prefetch_related("payment_plans")


def current_cashflow_receivable_items():
    return current_cashflow_items().filter(category=CashflowImportedItem.CATEGORY_RECEIVABLE)


def cashflow_items_missing_payment_plan():
    items = list(current_cashflow_outflow_items())
    return [item for item in items if not item.has_complete_payment_plan]


def contract_installment_effective_collection_date(installment):
    return installment.revised_collection_date or installment.expected_collection_date


def overdue_contract_installments(today=None):
    today = today or business_localdate()
    installments = (
        ContractCollectionInstallment.objects.select_related("contract_collection", "contract_collection__sales_manager")
        .filter(installment_amount__isnull=False)
        .order_by("contract_collection__company_name", "position")
    )
    overdue = []
    for installment in installments:
        if installment.is_collected:
            continue
        effective_date = contract_installment_effective_collection_date(installment)
        if effective_date and effective_date < today:
            overdue.append(installment)
    return overdue


def overdue_projected_collections(today=None):
    today = today or business_localdate()
    return list(
        CashflowProjectedCollection.objects.filter(
            Q(revised_collection_date__lt=today)
            | (Q(revised_collection_date__isnull=True) & Q(expected_collection_date__lt=today))
        ).order_by("company_name", "expected_collection_date")
    )


def cashflow_business_blockers(today=None):
    today = today or business_localdate()
    missing_plans = cashflow_items_missing_payment_plan()
    overdue_installments = overdue_contract_installments(today=today)
    overdue_projections = overdue_projected_collections(today=today)
    return {
        "missing_payment_plans": missing_plans,
        "overdue_installments": overdue_installments,
        "overdue_projected_collections": overdue_projections,
        "is_blocked": bool(missing_plans or overdue_installments or overdue_projections),
    }


def sync_cashflow_item_data(cashflow_item, cleaned_data):
    cashflow_item.primary_classification = cleaned_data["primary_classification"]
    cashflow_item.secondary_classification = cleaned_data["secondary_classification"]
    cashflow_item.cost_type = cleaned_data["cost_type"]
    cashflow_item.recurring_payable_day = cleaned_data["recurring_payable_day"] or None
    cashflow_item.save()
    cashflow_item.payment_plans.all().delete()
    CashflowPaymentPlanEntry.objects.bulk_create(
        [
            CashflowPaymentPlanEntry(
                cashflow_item=cashflow_item,
                amount=row["amount"],
                payment_date=row["payment_date"],
            )
            for row in cleaned_data["payment_plan_rows"]
        ]
    )
    return cashflow_item


def sync_projected_collection(projected_collection, cleaned_data):
    projected_collection.company_name = cleaned_data["company_name"]
    projected_collection.description = cleaned_data["description"]
    projected_collection.amount = cleaned_data["amount"]
    projected_collection.expected_collection_date = cleaned_data["expected_collection_date"]
    projected_collection.revised_collection_date = cleaned_data["revised_collection_date"]
    projected_collection.save()
    return projected_collection


def create_cashflow_work_order(cleaned_data, created_by):
    work_order = CashflowWorkOrderRequest.objects.create(
        description=cleaned_data["description"],
        party_name=cleaned_data["party_name"],
        total_amount=cleaned_data["total_amount"],
        created_by=created_by,
    )
    CashflowWorkOrderInstallment.objects.bulk_create(
        [
            CashflowWorkOrderInstallment(
                work_order=work_order,
                position=row["position"],
                amount=row["amount"],
                payment_date=row["payment_date"],
            )
            for row in cleaned_data["installment_rows"]
        ]
    )
    transaction.on_commit(lambda: send_cashflow_work_order_email(work_order.pk))
    return work_order


def send_cashflow_work_order_email(work_order_id):
    work_order = CashflowWorkOrderRequest.objects.prefetch_related("installments").select_related("created_by").get(
        pk=work_order_id
    )
    to_emails = unique_emails(active_finance_manager_emails())
    if not to_emails:
        raise ValueError("No active finance manager email is configured for work-order notifications.")
    html_lines = [
        f"<p>A business manager has requested a non-operational provision work order.</p>",
        f"<p><strong>Party:</strong> {work_order.party_name}<br>",
        f"<strong>Total amount:</strong> Rs {work_order.total_amount}<br>",
        f"<strong>Description:</strong> {work_order.description}<br>",
        f"<strong>Created by:</strong> {work_order.created_by.name}</p>",
        "<ul>",
    ]
    text_lines = [
        "A business manager has requested a non-operational provision work order.",
        f"Party: {work_order.party_name}",
        f"Total amount: Rs {work_order.total_amount}",
        f"Description: {work_order.description}",
        f"Created by: {work_order.created_by.name}",
        "Installments:",
    ]
    for installment in work_order.installments.all():
        html_lines.append(f"<li>Installment {installment.position}: Rs {installment.amount} on {installment.payment_date}</li>")
        text_lines.append(f"- Installment {installment.position}: Rs {installment.amount} on {installment.payment_date}")
    html_lines.append("</ul>")
    send_email(
        subject=f"Cashflow work order: {work_order.party_name}",
        html_body="".join(html_lines),
        text_body="\n".join(text_lines),
        to_emails=to_emails,
    )


def build_twelve_week_windows(start_date=None, weeks=12):
    start_date = start_date or business_localdate()
    windows = []
    for index in range(weeks):
        period_start = start_date + timedelta(days=index * 7)
        period_end = period_start + timedelta(days=6)
        windows.append(
            {
                "index": index + 1,
                "start_date": period_start,
                "end_date": period_end,
            }
        )
    return windows


def _future_month_dates(start_date, end_date):
    current_year = start_date.year
    current_month = start_date.month
    month_cursor = current_month + 1
    year_cursor = current_year
    dates = []
    while True:
        if month_cursor > 12:
            month_cursor = 1
            year_cursor += 1
        first_of_month = date(year_cursor, month_cursor, 1)
        if first_of_month > end_date:
            break
        dates.append(first_of_month)
        month_cursor += 1
    return dates


def build_cashflow_projection(start_date=None, weeks=12):
    start_date = start_date or business_localdate()
    windows = build_twelve_week_windows(start_date=start_date, weeks=weeks)
    end_date = windows[-1]["end_date"]
    outflow_rows = []
    inflow_rows = []

    for item in current_cashflow_outflow_items():
        for plan in item.payment_plans.all():
            if start_date <= plan.payment_date <= end_date:
                outflow_rows.append(
                    {
                        "source_type": item.category,
                        "label": item.party_name,
                        "description": item.description,
                        "amount": item.amount,
                        "transaction_amount": plan.amount,
                        "transaction_date": plan.payment_date,
                        "classification": item.primary_classification,
                        "secondary_classification": item.secondary_classification,
                        "object": item,
                    }
                )
        if item.cost_type == CashflowImportedItem.COST_RECURRING and item.recurring_payable_day:
            for first_of_month in _future_month_dates(start_date, end_date):
                day = min(item.recurring_payable_day, monthrange(first_of_month.year, first_of_month.month)[1])
                payment_date = date(first_of_month.year, first_of_month.month, day)
                if start_date <= payment_date <= end_date:
                    outflow_rows.append(
                        {
                            "source_type": "future_provision",
                            "label": item.party_name,
                            "description": item.description,
                            "amount": item.amount,
                            "transaction_amount": item.amount,
                            "transaction_date": payment_date,
                            "classification": item.primary_classification,
                            "secondary_classification": item.secondary_classification,
                            "object": item,
                        }
                    )

    installments = (
        ContractCollectionInstallment.objects.select_related("contract_collection", "contract_collection__sales_manager")
        .filter(installment_amount__isnull=False)
        .order_by("contract_collection__company_name", "position")
    )
    for installment in installments:
        if installment.is_collected:
            continue
        collection_date = contract_installment_effective_collection_date(installment)
        if collection_date and start_date <= collection_date <= end_date:
            inflow_rows.append(
                {
                    "source_type": "contract_installment",
                    "label": installment.contract_collection.company_name,
                    "description": f"Installment {installment.position}",
                    "transaction_amount": installment.installment_amount,
                    "transaction_date": collection_date,
                    "object": installment,
                }
            )

    for projected in CashflowProjectedCollection.objects.all().order_by("company_name", "expected_collection_date"):
        collection_date = projected.effective_collection_date
        if collection_date and start_date <= collection_date <= end_date:
            inflow_rows.append(
                {
                    "source_type": "projected_collection",
                    "label": projected.company_name,
                    "description": projected.description,
                    "transaction_amount": projected.amount,
                    "transaction_date": collection_date,
                    "object": projected,
                }
            )

    projection_rows = []
    for window in windows:
        week_inflows = [row for row in inflow_rows if window["start_date"] <= row["transaction_date"] <= window["end_date"]]
        week_outflows = [row for row in outflow_rows if window["start_date"] <= row["transaction_date"] <= window["end_date"]]
        inflow_total = sum((row["transaction_amount"] for row in week_inflows), Decimal("0.00"))
        outflow_total = sum((row["transaction_amount"] for row in week_outflows), Decimal("0.00"))
        projection_rows.append(
            {
                **window,
                "inflow_total": inflow_total,
                "outflow_total": outflow_total,
                "net_total": inflow_total - outflow_total,
                "inflows": week_inflows,
                "outflows": week_outflows,
            }
        )
    return {
        "start_date": start_date,
        "weeks": projection_rows,
    }


def sync_sales_conversation_data(sales_conversation, cleaned_data, uploaded_files=None, allow_company_name_edits=False):
    uploaded_files = uploaded_files or {}
    if not sales_conversation.pk or allow_company_name_edits:
        sales_conversation.company_name = cleaned_data["company_name"]
    sales_conversation.assigned_sales_manager = cleaned_data.get("assigned_sales_manager")
    sales_conversation.conversation_status = cleaned_data["conversation_status"]
    sales_conversation.proposal_status = cleaned_data["proposal_status"]
    sales_conversation.contract_signed = cleaned_data.get("contract_signed", False)
    sales_conversation.comments = cleaned_data.get("comments", "")
    sales_conversation.save()

    if allow_company_name_edits and sales_conversation.source_meeting_id:
        prospect = sales_conversation.source_meeting.prospect
        if prospect.company_name != sales_conversation.company_name:
            prospect.company_name = sales_conversation.company_name
            prospect.save(update_fields=["company_name", "updated_at"])
    if allow_company_name_edits and hasattr(sales_conversation, "contract_collection"):
        contract_collection = sales_conversation.contract_collection
        if contract_collection.company_name != sales_conversation.company_name:
            contract_collection.company_name = sales_conversation.company_name
            contract_collection.save(update_fields=["company_name", "updated_at"])

    sales_conversation.contacts.all().delete()
    SalesConversationContact.objects.bulk_create(
        [
            SalesConversationContact(
                sales_conversation=sales_conversation,
                position=row["position"],
                name=row["name"],
                email=row["email"],
                whatsapp_number=row["whatsapp_number"],
            )
            for row in cleaned_data["contact_rows"]
        ]
    )

    sales_conversation.brands.all().delete()
    SalesConversationBrand.objects.bulk_create(
        [
            SalesConversationBrand(sales_conversation=sales_conversation, name=brand_name)
            for brand_name in cleaned_data["brands_input"]
        ]
    )

    for category, field_name in (
        (SalesConversationFile.CATEGORY_SOLUTION, "solution_files"),
        (SalesConversationFile.CATEGORY_PROPOSAL, "proposal_files"),
    ):
        for uploaded_file in uploaded_files.get(field_name, []):
            SalesConversationFile.objects.create(
                sales_conversation=sales_conversation,
                category=category,
                file=uploaded_file,
            )
    return sales_conversation


def get_or_create_sales_conversation_from_meeting(meeting, created_by=None):
    defaults = {
        "company_name": meeting.prospect.company_name,
        "assigned_sales_manager": default_sales_manager(),
        "created_by": created_by or meeting.scheduled_by,
    }
    sales_conversation, created = SalesConversation.objects.get_or_create(
        source_meeting=meeting,
        defaults=defaults,
    )
    if created:
        SalesConversationContact.objects.create(
            sales_conversation=sales_conversation,
            position=1,
            name=meeting.prospect.contact_name,
            email=meeting.prospect_email or meeting.prospect.prospect_email,
            whatsapp_number="",
        )
    elif not sales_conversation.assigned_sales_manager_id:
        manager = default_sales_manager()
        if manager:
            sales_conversation.assigned_sales_manager = manager
            sales_conversation.save(update_fields=["assigned_sales_manager", "updated_at"])
    return sales_conversation


def backfill_sales_conversations_from_happened_meetings():
    meetings = list(
        Meeting.objects.filter(status=Meeting.STATUS_HAPPENED, sales_conversation__isnull=True)
        .select_related("prospect", "scheduled_by")
        .order_by("pk")
    )
    for meeting in meetings:
        get_or_create_sales_conversation_from_meeting(meeting, created_by=meeting.scheduled_by)
    return len(meetings)


def get_or_create_contract_collection_from_sales_conversation(sales_conversation, created_by=None):
    defaults = {
        "company_name": sales_conversation.company_name,
        "sales_manager": sales_conversation.assigned_sales_manager,
        "created_by": created_by or sales_conversation.created_by,
    }
    contract_collection, created = ContractCollection.objects.get_or_create(
        source_sales_conversation=sales_conversation,
        defaults=defaults,
    )
    if created:
        ContractCollectionContact.objects.bulk_create(
            [
                ContractCollectionContact(
                    contract_collection=contract_collection,
                    position=contact.position,
                    name=contact.name,
                    email=contact.email,
                    whatsapp_number=contact.whatsapp_number,
                )
                for contact in sales_conversation.contacts.all()
            ]
        )
    else:
        updates = []
        if not contract_collection.sales_manager_id and sales_conversation.assigned_sales_manager_id:
            contract_collection.sales_manager = sales_conversation.assigned_sales_manager
            updates.append("sales_manager")
        if updates:
            updates.append("updated_at")
            contract_collection.save(update_fields=updates)
    return contract_collection


def sync_contract_collection_data(
    contract_collection,
    cleaned_data,
    uploaded_files=None,
    allow_locked_field_edits=False,
    allow_expected_collection_date_edits=False,
):
    uploaded_files = uploaded_files or {}
    contract_collection.sales_manager = cleaned_data.get("sales_manager")
    if allow_locked_field_edits or contract_collection.contract_value is None:
        contract_collection.contract_value = cleaned_data.get("contract_value")
    contract_collection.save()
    today = business_localdate()
    installments_requiring_immediate_notification = []

    contract_collection.contacts.all().delete()
    ContractCollectionContact.objects.bulk_create(
        [
            ContractCollectionContact(
                contract_collection=contract_collection,
                position=row["position"],
                name=row["name"],
                email=row["email"],
                whatsapp_number=row["whatsapp_number"],
            )
            for row in cleaned_data["contact_rows"]
        ]
    )

    if uploaded_files.get("contract_files") and not contract_collection.files.exists():
        for uploaded_file in uploaded_files["contract_files"]:
            ContractCollectionFile.objects.create(
                contract_collection=contract_collection,
                file=uploaded_file,
            )

    installments_by_position = {item.position: item for item in contract_collection.installments.all()}
    for row in cleaned_data["installment_rows"]:
        installment = installments_by_position.get(row["position"])
        if not installment:
            installment = ContractCollectionInstallment.objects.create(
                contract_collection=contract_collection,
                position=row["position"],
            )
        if (allow_locked_field_edits or installment.installment_amount is None) and row["installment_amount"] is not None:
            installment.installment_amount = row["installment_amount"]
        if row["invoice_date"] is not None and (
            allow_locked_field_edits or installment.invoice_date is None
        ):
            if installment.invoice_date != row["invoice_date"]:
                installment.invoice_notification_sent_at = None
            installment.invoice_date = row["invoice_date"]
        if row["expected_collection_date"] is not None and (
            allow_locked_field_edits
            or allow_expected_collection_date_edits
            or installment.expected_collection_date is None
        ):
            installment.expected_collection_date = row["expected_collection_date"]
        installment.revised_collection_date = row["revised_collection_date"]
        installment.contract_summary = row["contract_summary"]
        installment.invoiced_service_description = row["invoiced_service_description"]
        installment.legal_due_reason = row["legal_due_reason"]
        installment.save()
        if installment.invoice_date == today and installment.invoice_notification_sent_at is None:
            installments_requiring_immediate_notification.append(installment.pk)

    if contract_collection.source_sales_conversation_id and contract_collection.sales_manager_id:
        sales_conversation = contract_collection.source_sales_conversation
        if sales_conversation.assigned_sales_manager_id != contract_collection.sales_manager_id:
            sales_conversation.assigned_sales_manager = contract_collection.sales_manager
            sales_conversation.save(update_fields=["assigned_sales_manager", "updated_at"])
    if installments_requiring_immediate_notification:
        transaction.on_commit(
            lambda: _send_immediate_invoice_due_notifications_after_commit(
                installments_requiring_immediate_notification
            )
        )
    return contract_collection


def sync_finance_collection_data(contract_collection, cleaned_data):
    installments_by_position = {item.position: item for item in contract_collection.installments.all()}
    for row in cleaned_data["finance_rows"]:
        installment = installments_by_position.get(row["position"])
        if not installment:
            installment = ContractCollectionInstallment.objects.create(
                contract_collection=contract_collection,
                position=row["position"],
            )
        installment.collected_amount = row["collected_amount"]
        installment.collection_date = row["collection_date"]
        installment.save()
    return contract_collection


def build_pending_collections(contract_query_set, today=None):
    today = today or business_localdate()
    installments = ContractCollectionInstallment.objects.select_related(
        "contract_collection",
        "contract_collection__sales_manager",
    ).prefetch_related("contract_collection__contacts")
    installments = installments.filter(contract_collection__in=contract_query_set)

    invoiced = installments.filter(invoice_date__lt=today).filter(
        Q(collection_date__isnull=True)
        | Q(collected_amount__isnull=True)
        | Q(collected_amount__lt=F("installment_amount"))
    )
    to_be_invoiced = installments.filter(
        Q(invoice_date__isnull=True) | Q(invoice_date__gte=today)
    ).filter(installment_amount__isnull=False)

    return {
        "invoiced_pending": invoiced.order_by("invoice_date", "contract_collection__company_name", "position"),
        "invoiced_pending_total": sum(
            [item.installment_amount or Decimal("0.00") for item in invoiced],
            Decimal("0.00"),
        ),
        "yet_to_invoice": to_be_invoiced.order_by("invoice_date", "contract_collection__company_name", "position"),
        "yet_to_invoice_total": sum(
            [item.installment_amount or Decimal("0.00") for item in to_be_invoiced],
            Decimal("0.00"),
        ),
    }


def send_invoice_due_email(installment):
    contract = installment.contract_collection
    settings_obj = SystemSetting.load()
    html_body = render_to_string(
        "emails/invoice_due.html",
        {"contract": contract, "installment": installment, "settings_obj": settings_obj},
    )
    text_body = render_to_string(
        "emails/invoice_due.txt",
        {"contract": contract, "installment": installment, "settings_obj": settings_obj},
    )
    to_emails = unique_emails(active_finance_manager_emails())
    cc_emails = unique_emails(
        active_leadgen_supervisor_emails(),
        active_sales_manager_emails(),
    )
    if not to_emails:
        raise ValueError("No active finance manager email is configured for invoice notifications.")
    send_email(
        subject=f"Invoice to be raised today: {contract.company_name} / installment {installment.position}",
        html_body=html_body,
        text_body=text_body,
        to_emails=to_emails,
        cc_emails=cc_emails,
    )


def send_due_invoice_notifications(target_date=None, now=None):
    target_date = target_date or business_localdate(now=now)
    installments = ContractCollectionInstallment.objects.select_related(
        "contract_collection",
        "contract_collection__sales_manager",
    ).filter(
        invoice_date=target_date,
        invoice_notification_sent_at__isnull=True,
    )
    sent_count = 0
    for installment in installments:
        try:
            send_invoice_due_email(installment)
        except Exception:
            logger.exception(
                "Failed to send invoice due email for contract_collection_id=%s installment=%s",
                installment.contract_collection.contract_collection_id,
                installment.position,
            )
            continue
        installment.invoice_notification_sent_at = timezone.now()
        installment.save(update_fields=["invoice_notification_sent_at"])
        sent_count += 1
    return sent_count


def _send_immediate_invoice_due_notifications_after_commit(installment_ids):
    try:
        today = business_localdate()
        installments = ContractCollectionInstallment.objects.select_related(
            "contract_collection",
            "contract_collection__sales_manager",
        ).filter(
            pk__in=installment_ids,
            invoice_date=today,
            invoice_notification_sent_at__isnull=True,
        )
        for installment in installments:
            try:
                send_invoice_due_email(installment)
            except Exception:
                logger.exception(
                    "Failed to send immediate invoice due email for contract_collection_id=%s installment=%s",
                    installment.contract_collection.contract_collection_id,
                    installment.position,
                )
                continue
            installment.invoice_notification_sent_at = timezone.now()
            installment.save(update_fields=["invoice_notification_sent_at"])
    except Exception:
        logger.exception(
            "Failed to process immediate invoice due notifications for installment_ids=%s",
            installment_ids,
        )


def _send_meeting_invitation_after_commit(meeting_id):
    try:
        meeting = Meeting.objects.select_related("prospect", "scheduled_by").get(pk=meeting_id)
        send_meeting_invitation(meeting)
    except Exception:
        logger.exception("Failed to send meeting invitation for meeting_id=%s", meeting_id)


def _send_did_not_happen_email_after_commit(meeting_id):
    try:
        meeting = Meeting.objects.select_related("prospect", "scheduled_by").get(pk=meeting_id)
        send_did_not_happen_email(meeting)
    except Exception:
        logger.exception("Failed to send meeting follow-up email for meeting_id=%s", meeting_id)


@transaction.atomic
def apply_call_outcome(prospect, staff, cleaned_data):
    settings_obj = SystemSetting.load()
    outcome = cleaned_data["outcome"]
    reason = cleaned_data.get("reason", "")
    follow_up_date = cleaned_data.get("follow_up_date")
    scheduled_for = cleaned_data.get("scheduled_for")
    prospect_email = cleaned_data.get("prospect_email", "")
    meeting_platform = cleaned_data.get("meeting_platform", "")
    meeting = None

    if outcome == ProspectStatusUpdate.OUTCOME_FOLLOW_UP:
        prospect.workflow_status = Prospect.WORKFLOW_FOLLOW_UP
        prospect.follow_up_date = follow_up_date
        prospect.follow_up_reason = reason
        prospect.decline_reason = ""
    elif outcome == ProspectStatusUpdate.OUTCOME_DECLINED:
        prospect.workflow_status = Prospect.WORKFLOW_DECLINED
        prospect.decline_reason = reason
        prospect.follow_up_date = None
        prospect.follow_up_reason = ""
    elif outcome == ProspectStatusUpdate.OUTCOME_SCHEDULED:
        if timezone.is_naive(scheduled_for):
            scheduled_for = timezone.make_aware(scheduled_for, ZoneInfo(settings_obj.default_timezone))
        recipient_emails = unique_emails(
            [prospect_email, settings_obj.supervisor_sender_email, staff.email],
            active_sales_manager_emails(),
            settings_obj.sales_emails(),
        )
        meeting = Meeting.objects.create(
            prospect=prospect,
            scheduled_by=staff,
            scheduled_for=scheduled_for,
            prospect_email=prospect_email,
            meeting_platform=meeting_platform,
            recipient_emails=recipient_emails,
        )
        prospect.workflow_status = Prospect.WORKFLOW_SCHEDULED
        prospect.prospect_email = prospect_email
        prospect.follow_up_date = None
        prospect.follow_up_reason = ""
        prospect.decline_reason = ""

    prospect.save()
    ProspectStatusUpdate.objects.create(
        prospect=prospect,
        staff=staff,
        outcome=outcome,
        reason=reason,
        follow_up_date=follow_up_date,
        scheduled_for=scheduled_for,
        prospect_email=prospect_email,
    )
    if meeting:
        transaction.on_commit(lambda: _send_meeting_invitation_after_commit(meeting.pk))
        transaction.on_commit(lambda: _send_immediate_day_before_reminder_after_commit(meeting.pk))
    return meeting


@transaction.atomic
def update_meeting_outcome(meeting, status, updated_by=None):
    meeting.status = status
    meeting.outcome_updated_at = timezone.now()
    meeting.save(update_fields=["status", "outcome_updated_at", "updated_at"])

    if status == Meeting.STATUS_HAPPENED:
        meeting.prospect.workflow_status = Prospect.WORKFLOW_MEETING_HAPPENED
        meeting.prospect.save(update_fields=["workflow_status", "updated_at"])
        ProspectStatusUpdate.objects.create(
            prospect=meeting.prospect,
            staff=meeting.scheduled_by,
            outcome=ProspectStatusUpdate.OUTCOME_MEETING_HAPPENED,
            scheduled_for=meeting.scheduled_for,
            prospect_email=meeting.prospect_email,
        )
        get_or_create_sales_conversation_from_meeting(meeting, created_by=updated_by)
        return

    meeting.prospect.workflow_status = Prospect.WORKFLOW_FOLLOW_UP
    meeting.prospect.follow_up_date = meeting.scheduled_for.astimezone(
        ZoneInfo(SystemSetting.load().default_timezone)
    ).date()
    outcome_reason = "Meeting did not happen - no show" if status == Meeting.STATUS_NO_SHOW else "Meeting did not happen"
    meeting.prospect.follow_up_reason = outcome_reason
    meeting.prospect.save(update_fields=["workflow_status", "follow_up_date", "follow_up_reason", "updated_at"])
    ProspectStatusUpdate.objects.create(
        prospect=meeting.prospect,
        staff=meeting.scheduled_by,
        outcome=ProspectStatusUpdate.OUTCOME_MEETING_DID_NOT_HAPPEN,
        reason=outcome_reason,
        follow_up_date=meeting.prospect.follow_up_date,
        scheduled_for=meeting.scheduled_for,
        prospect_email=meeting.prospect_email,
    )
    transaction.on_commit(lambda: _send_did_not_happen_email_after_commit(meeting.pk))


@transaction.atomic
def reschedule_meeting(meeting, rescheduled_for, updated_by=None):
    settings_obj = SystemSetting.load()
    if timezone.is_naive(rescheduled_for):
        rescheduled_for = timezone.make_aware(rescheduled_for, ZoneInfo(settings_obj.default_timezone))

    meeting.status = Meeting.STATUS_RESCHEDULED
    meeting.outcome_updated_at = timezone.now()
    meeting.save(update_fields=["status", "outcome_updated_at", "updated_at"])

    new_meeting = Meeting.objects.create(
        prospect=meeting.prospect,
        scheduled_by=meeting.scheduled_by,
        scheduled_for=rescheduled_for,
        prospect_email=meeting.prospect_email,
        meeting_platform=meeting.meeting_platform,
        recipient_emails=meeting.recipient_emails,
    )

    meeting.prospect.workflow_status = Prospect.WORKFLOW_SCHEDULED
    meeting.prospect.prospect_email = meeting.prospect_email
    meeting.prospect.follow_up_date = None
    meeting.prospect.follow_up_reason = ""
    meeting.prospect.save(
        update_fields=["workflow_status", "prospect_email", "follow_up_date", "follow_up_reason", "updated_at"]
    )

    ProspectStatusUpdate.objects.create(
        prospect=meeting.prospect,
        staff=meeting.scheduled_by,
        outcome=ProspectStatusUpdate.OUTCOME_SCHEDULED,
        reason=MEETING_RESCHEDULE_REASON,
        scheduled_for=rescheduled_for,
        prospect_email=meeting.prospect_email,
    )
    transaction.on_commit(lambda: _send_meeting_invitation_after_commit(new_meeting.pk))
    transaction.on_commit(lambda: _send_immediate_day_before_reminder_after_commit(new_meeting.pk))
    return new_meeting


@transaction.atomic
def delete_meeting(meeting):
    prospect = meeting.prospect
    scheduled_for_date = meeting.scheduled_for.astimezone(ZoneInfo(SystemSetting.load().default_timezone)).date()
    deleted_was_scheduled = meeting.status == Meeting.STATUS_SCHEDULED
    scheduled_by = meeting.scheduled_by
    prospect_email = meeting.prospect_email

    meeting.delete()

    if deleted_was_scheduled and prospect.workflow_status == Prospect.WORKFLOW_SCHEDULED:
        has_remaining_scheduled = prospect.meetings.filter(status=Meeting.STATUS_SCHEDULED).exists()
        if not has_remaining_scheduled:
            prospect.workflow_status = Prospect.WORKFLOW_FOLLOW_UP
            prospect.follow_up_date = scheduled_for_date
            prospect.follow_up_reason = "Meeting deleted by supervisor"
            prospect.save(update_fields=["workflow_status", "follow_up_date", "follow_up_reason", "updated_at"])
            ProspectStatusUpdate.objects.create(
                prospect=prospect,
                staff=scheduled_by,
                outcome=ProspectStatusUpdate.OUTCOME_FOLLOW_UP,
                reason="Meeting deleted by supervisor",
                follow_up_date=scheduled_for_date,
                prospect_email=prospect_email,
            )
